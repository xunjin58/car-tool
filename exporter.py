from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADERS = ["品牌", "车型", "配置版本", "颜色", "价格", "数量", "地区", "联系方式", "备注"]

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def export_to_excel(records: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "汽车报价"

    # header row
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22

    # data rows
    for row_idx, record in enumerate(records, 2):
        fill = ALT_FILL if row_idx % 2 == 0 else None
        for col, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col, value=record.get(header, ""))
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_idx].height = 18

    # auto column width
    for col in range(1, len(HEADERS) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 30)

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
